from rich.console import Console
from rich.theme import Theme
from rich.table import Table
from os import mkdir, listdir
from os.path import exists
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

semantic = Theme({
    'prompt': '#10a4e8 bold',
    'info': '#edae1a bold',
    'error': '#e82610',
    'success': '#10e83f',
    'dim': '#404040'
})

console = Console(theme=semantic, highlight=False)

def hr():
    console.print('')
    console.rule(style='dim')
    console.print('')

record = []

def init():
    hr()
    console.print('GPACALC', style='info')
    console.print('CREATED BY KEVIN QU', style='info')
    menu()

def menu():
    hr()
    console.print('What would you like to do?', style='prompt')
    console.print('[ n ] create new record')
    console.print('[ v ] view current record')
    console.print('[ d ] delete record element')
    console.print('[ e ] export record as spreadsheet')
    console.print('[ h ] help')
    console.print('[ x ] exit')
    inpt = console.input('\n> ').lower()

    if inpt == 'n':
        new()
    elif inpt == 'v':
        view()
    elif inpt == 'd':
        delete()
    elif inpt == 'e':
        export()
    elif inpt == 'h':
        hr()
        console.print('[ c ] create new record - creates a blank record and prompts you for element values')
        console.print('[ v ] view current record - view current record')
        console.print('[ d ] delete record element - delete a specific class in the record')
        console.print('[ e ] export record as spreadsheet - export current record as an excel (.xlsx) file')
        console.input('\npress [ enter ] to continue > ')
    elif inpt == 'x':
        console.print('\nThank you for using gpacalc. Exiting program.', style='info')
        exit()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')

    menu()

def new():
    global record

    def create():
        global record

        inpt = ''
        while inpt.lower() != 'done':
            console.print('\nPlease enter a class | letter grade | weight', style='prompt')
            console.print("Enter 'done' to finish", style='prompt')
            inpt = console.input('> ')

            if inpt != 'done':
                try:
                    inpt_split = inpt.split(' | ')
                    inpt_split[1] = inpt_split[1].lower()
                    inpt_split[2] = inpt_split[2].lower()

                    if inpt_split[0] and len(inpt_split[0]) <= 24 and inpt_split[0] not in record_classes() and inpt_split[1] in ['a', 'a-', 'b+', 'b', 'b-', 'c+', 'c', 'c-', 'd+', 'd', 'd-', 'f'] and inpt_split[2] in ['r', 'p', 'f'] and len(inpt_split) <= 3:
                        gpa = 0
                        if inpt_split[1] == 'a' and inpt_split[2] == 'f':
                            gpa = 5.0
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'f':
                            gpa = 4.667
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'f':
                            gpa = 4.333
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'f':
                            gpa = 4.0
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'f':
                            gpa = 3.667
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'f':
                            gpa = 3.333
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'f':
                            gpa = 3.0
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'f':
                            gpa = 2.667
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'f':
                            gpa = 2.333
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'f':
                            gpa = 2.0
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'f':
                            gpa = 1.667
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'f':
                            gpa = 0.0
                        elif inpt_split[1] == 'a' and inpt_split[2] == 'p':
                            gpa = 4.5
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'p':
                            gpa = 4.167
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'p':
                            gpa = 3.833
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'p':
                            gpa = 3.5
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'p':
                            gpa = 3.167
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'p':
                            gpa = 2.833
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'p':
                            gpa = 2.5
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'p':
                            gpa = 2.167
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'p':
                            gpa = 1.833
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'p':
                            gpa = 1.5
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'p':
                            gpa = 1.167
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'p':
                            gpa = 0.0
                        elif inpt_split[1] == 'a' and inpt_split[2] == 'r':
                            gpa = 4.0
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'r':
                            gpa = 3.667
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'r':
                            gpa = 3.333
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'r':
                            gpa = 3.0
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'r':
                            gpa = 2.667
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'r':
                            gpa = 2.333
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'r':
                            gpa = 2.0
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'r':
                            gpa = 1.667
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'r':
                            gpa = 1.333
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'r':
                            gpa = 1.0
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'r':
                            gpa = 0.667
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'r':
                            gpa = 0.0
                        record.append({'class': inpt_split[0], 'grade': inpt_split[1], 'weight': inpt_split[2], 'gpa': gpa})
                        console.print('\nClass added to record', style='success')
                    else:
                        console.print('\n[ERROR] Invalid input syntax. Please reference the tutorial if needed.', style='error')
                        console.input('press [ enter ] to continue > ')
                except IndexError:
                    console.print('\n[ERROR] Invalid input syntax. Please reference the tutorial if needed.', style='error')
                    console.input('press [ enter ] to continue > ')

        console.print('\nNew record saved.', style='success')
        console.input('press [ enter ] to continue > ')

    hr()

    console.print('Would you like to view the tutorial?', style='prompt')
    console.print('[ y ] yes')
    console.print('[ n ] no, i know what to do already')
    inpt = console.input('\n> ').lower()

    if inpt == 'y':
        console.print('\nEnter the names of classes in the following format: class | letter grade | weight')
        console.print('The class name can be anything 24 characters or under and cannot be used twice')
        console.print('The letter grade is your letter grade in the class (pluses and minuses are supported, but a+ is not)')
        console.print('The weight can be entered in as either \'r\' (regular), \'p\' (partial), or \'f\' (full)')
        console.print('Example: computer science | a- | f')
        console.print('At any time, enter \'done\' to finish and save the classes')
        console.input('\npress [ enter ] to continue > ')
        create()
    elif inpt == 'n':
        create()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')
        new()

def view():
    hr()

    if len(record) > 0:
        table = Table(show_header=True)
        table.add_column('CLASS', width=24)
        table.add_column('LG', width=2)
        table.add_column('W', width=1)
        table.add_column('GPA', width=5)

        for el in record:
            table.add_row(f'{el["class"]}', f'{el["grade"].upper()}', f'{el["weight"].upper()}', f'{el["gpa"]}')

        console.print(table)

        console.print(f'Cumulative Weighted GPA: [prompt]{calculateGPA("w")}[/prompt]')
        console.print(f'Cumulative Unweighted GPA: [prompt]{calculateGPA("u")}[/prompt]')
    else:
        console.print('Record is empty. Try adding classes to it by entering [ n ] in the menu.', style='error')

    console.input('\npress [ enter ] to continue > ')

def delete():
    global record
    hr()
    console.print('What class would you like to delete (note that this is case-sensitive)?', style='prompt')
    console.print('Enter [ x ] to cancel', style='prompt')
    del_class = console.input('\n> ')

    if del_class.lower() == 'x':
        menu()
    elif del_class in record_classes():
        ls = record_classes()
        idx = 0
        itr = 0
        while itr < len(ls):
            if ls[itr] == del_class:
                idx = itr
            itr += 1

        record.pop(idx)

        console.print('\nClass successfuly deleted!', style='success')
        console.input('press [ enter ] to continue > ')
        menu()
    else:
        console.print('\n[ERROR] No matching class name. Please try again.', style='error')
        console.input('press [ enter ] to continue > ')
        delete()

def export():
    hr()
    console.print("Would you like to export your current record to a Excel (.xlsx) file?", style='prompt')
    console.print('[ y ] yes')
    console.print('[ n ] no')

    inpt = console.input('\n> ').lower()

    if inpt == 'y' and len(record) > 0:
        files = listdir()
        if 'records' not in files:
            mkdir('records')

        if exists('records/gpacalc.xlsx'):
            console.print("\nYou currently have another file titled 'gpacalc.xlsx'", style='prompt')
            console.print('Would you like to overwrite this file?', style='prompt')
            console.print('[ y ] yes, overwrite the file')
            console.print('[ n ] no, take me back to the menu')
            inpt = console.input('\n> ').lower()

            if inpt == 'y':
                generateSheet()
            elif inpt == 'n':
                menu()
            else:
                console.print('\n[ERROR] Please enter a valid option.', style='error')
                console.input('press [ enter ] to continue > ')
                export()
        else:
            generateSheet()
    elif inpt == 'y' and len(record) == 0:
        console.print('\n[ERROR] You must have at least one element in your record to export it as .csv', style='error')
        console.input('press [ enter ] to continue > ')
    elif inpt == 'n':
        menu()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')

def calculateGPA(weight):
    addGPA = 0
    if weight == 'w': # weighted
        for el in record:
            addGPA += el['gpa']
    elif weight == 'u': # unweighted
        for el in record:
            if el['grade'] == 'a':
                addGPA += 4.0
            elif el['grade'] == 'a-':
                addGPA += 3.667
            elif el['grade'] == 'b+':
                addGPA += 3.333
            elif el['grade'] == 'b':
                addGPA += 3.0
            elif el['grade'] == 'b-':
                addGPA += 2.667
            elif el['grade'] == 'c+':
                addGPA += 2.333
            elif el['grade'] == 'c':
                addGPA += 2.0
            elif el['grade'] == 'c-':
                addGPA += 1.667
            elif el['grade'] == 'd+':
                addGPA += 1.333
            elif el['grade'] == 'd':
                addGPA += 1.0
            elif el['grade'] == 'd-':
                addGPA += 0.667
            elif el['grade'] == 'f':
                addGPA += 0.0
    return round(addGPA / len(record), 3)

def generateSheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'gpacalc'

    worksheet.cell(row=1, column=1, value='CLASS')
    worksheet.cell(row=1, column=2, value='GRADE')
    worksheet.cell(row=1, column=3, value='WEIGHT')
    worksheet.cell(row=1, column=4, value='GPA')
    worksheet.cell(row=3, column=1, value='Weighted GPA')
    worksheet.cell(row=3, column=4, value=f"{calculateGPA('w')}")
    worksheet.cell(row=4, column=1, value='Unweighted GPA')
    worksheet.cell(row=4, column=4, value=f"{calculateGPA('u')}")

    worksheet['A1'].font = Font(bold=True)
    worksheet['B1'].font = Font(bold=True)
    worksheet['C1'].font = Font(bold=True)
    worksheet['D1'].font = Font(bold=True)

    worksheet['D3'].alignment = Alignment(horizontal='right')
    worksheet['D4'].alignment = Alignment(horizontal='right')

    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 10

    itr = 0
    while itr < len(record):
        worksheet.cell(row=itr+6, column=1, value=record[itr]['class'])
        worksheet.cell(row=itr+6, column=2, value=record[itr]['grade'].upper())
        worksheet.cell(row=itr+6, column=3, value=record[itr]['weight'].upper())
        worksheet.cell(row=itr+6, column=4, value=record[itr]['gpa'])
        itr += 1

    workbook.save(filename='records/gpacalc.xlsx')

    console.print("\nSpreadsheet titled 'gpacalc.xlsx' successfully generated in the records folder!", style='success')
    console.print('To open this file, you should use an external tool such as Google Sheets or Microsoft Excel.')
    console.input('press [ enter ] to continue > ')
    menu()

def record_classes():
    class_names = []
    for el in record:
        class_names.append(f'{el["class"]}')
    return class_names

init()
