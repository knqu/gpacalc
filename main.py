from rich.console import Console
from rich.theme import Theme
from rich.table import Table
from os import mkdir, listdir
from os.path import exists
from openpyxl import Workbook

semantic = Theme({
    'prompt': '#10a4e8 bold',
    'info': '#edae1a bold',
    'error': '#e82610',
    'success': '#10e83f',
    'dim': '#404040'
})

console = Console(theme=semantic, highlight=False)

def hr():
    console.print('')
    console.rule(style='dim')
    console.print('')

record = []

def init():
    hr()
    console.print('GPACALC', style='info')
    console.print('CREATED BY KEVIN QU', style='info')
    menu()

def menu():
    hr()
    console.print('What would you like to do?', style='prompt')
    console.print('[ n ] create new record')
    console.print('[ v ] view current record')
    console.print('[ e ] edit current record')
    console.print('[ s ] export record as spreadsheet')
    console.print('[ h ] help')
    console.print('[ x ] exit')
    inpt = console.input('\n> ')

    if inpt == 'n':
        new()
    elif inpt == 'v':
        view()
    elif inpt == 'e':
        edit()
    elif inpt == 's':
        export()
    elif inpt == 'h':
        hr()
        console.print('[ c ] create new record - creates a blank record and prompts you for element values')
        console.print('[ v ] view current record - view current record')
        console.print('[ e ] edit current record - change elements of current record')
        console.print('[ s ] export record as spreadsheet - export current record as an excel (.xlsx) file')
        console.input('\npress [ enter ] to continue > ')
    elif inpt == 'x':
        console.print('\nThank you for using gpacalc. Exiting program.', style='info')
        exit()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')

    menu()

def new():
    global record

    def create():
        global record

        inpt = ''
        while inpt != 'done':
            console.print('\nPlease enter a class | letter grade | weight', style='prompt')
            inpt = console.input('> ')

            if inpt != 'done':
                try:
                    inpt_split = inpt.split(' | ')
                    inpt_split[1] = inpt_split[1].lower()
                    inpt_split[2] = inpt_split[2].lower()

                    if inpt_split[0] and len(inpt_split[0]) <= 24 and inpt_split[0] not in record_classes() and inpt_split[1] in ['a', 'a-', 'b+', 'b', 'b-', 'c+', 'c', 'c-', 'd+', 'd', 'd-', 'f'] and inpt_split[2] in ['r', 'p', 'f'] and len(inpt_split) <= 3:
                        gpa = 0
                        if inpt_split[1] == 'a' and inpt_split[2] == 'f':
                            gpa = 5.0
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'f':
                            gpa = 4.667
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'f':
                            gpa = 4.333
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'f':
                            gpa = 4.0
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'f':
                            gpa = 3.667
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'f':
                            gpa = 3.333
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'f':
                            gpa = 3.0
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'f':
                            gpa = 2.667
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'f':
                            gpa = 2.333
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'f':
                            gpa = 2.0
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'f':
                            gpa = 1.667
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'f':
                            gpa = 0.0
                        elif inpt_split[1] == 'a' and inpt_split[2] == 'p':
                            gpa = 4.5
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'p':
                            gpa = 4.167
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'p':
                            gpa = 3.833
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'p':
                            gpa = 3.5
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'p':
                            gpa = 3.167
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'p':
                            gpa = 2.833
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'p':
                            gpa = 2.5
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'p':
                            gpa = 2.167
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'p':
                            gpa = 1.833
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'p':
                            gpa = 1.5
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'p':
                            gpa = 1.167
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'p':
                            gpa = 0.0
                        elif inpt_split[1] == 'a' and inpt_split[2] == 'r':
                            gpa = 4.0
                        elif inpt_split[1] == 'a-' and inpt_split[2] == 'r':
                            gpa = 3.667
                        elif inpt_split[1] == 'b+' and inpt_split[2] == 'r':
                            gpa = 3.333
                        elif inpt_split[1] == 'b' and inpt_split[2] == 'r':
                            gpa = 3.0
                        elif inpt_split[1] == 'b-' and inpt_split[2] == 'r':
                            gpa = 2.667
                        elif inpt_split[1] == 'c+' and inpt_split[2] == 'r':
                            gpa = 2.333
                        elif inpt_split[1] == 'c' and inpt_split[2] == 'r':
                            gpa = 2.0
                        elif inpt_split[1] == 'c-' and inpt_split[2] == 'r':
                            gpa = 1.667
                        elif inpt_split[1] == 'd+' and inpt_split[2] == 'r':
                            gpa = 1.333
                        elif inpt_split[1] == 'd' and inpt_split[2] == 'r':
                            gpa = 1.0
                        elif inpt_split[1] == 'd-' and inpt_split[2] == 'r':
                            gpa = 0.667
                        elif inpt_split[1] == 'f' and inpt_split[2] == 'r':
                            gpa = 0.0
                        record.append({'class': inpt_split[0], 'grade': inpt_split[1], 'weight': inpt_split[2], 'gpa': gpa})
                        console.print('\nClass added to record', style='success')
                    else:
                        console.print('\n[ERROR] Invalid input syntax. Please reference the tutorial if needed.', style='error')
                        console.input('press [ enter ] to continue > ')
                except IndexError:
                    console.print('\n[ERROR] Invalid input syntax. Please reference the tutorial if needed.', style='error')
                    console.input('press [ enter ] to continue > ')

        console.print('\nNew record saved.', style='success')
        console.input('press [ enter ] to continue > ')

    hr()

    console.print('Would you like to view the tutorial?', style='prompt')
    console.print('[ y ] yes')
    console.print('[ n ] no, i know what to do already')
    inpt = console.input('\n> ')

    if inpt == 'y':
        console.print('\nEnter the names of classes in the following format: class | letter grade | weight')
        console.print('The class name can be anything 24 characters or under and cannot be used twice')
        console.print('The letter grade is your letter grade in the class (pluses and minuses are supported, but a+ is not)')
        console.print('The weight can be entered in as either \'r\' (regular), \'p\' (partial), or \'f\' (full)')
        console.print('Example: computer science | a- | f')
        console.print('At any time, enter \'done\' to finish and save the classes')
        console.input('\npress [ enter ] to continue > ')
        create()
    elif inpt == 'n':
        create()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')
        new()

def view():
    hr()

    if len(record) > 0:
        table = Table(show_header=True)
        table.add_column('CLASS', width=24)
        table.add_column('LG', width=2)
        table.add_column('W', width=1)
        table.add_column('GPA', width=5)

        for el in record:
            table.add_row(f'{el["class"]}', f'{el["grade"].upper()}', f'{el["weight"].upper()}', f'{el["gpa"]}')

        console.print(table)

        console.print(f'Cumulative Weighted GPA: [prompt]{calculateGPA("w")}[/prompt]')
        console.print(f'Cumulative Unweighted GPA: [prompt]{calculateGPA("u")}[/prompt]')
    else:
        console.print('Record is empty. Try adding classes to it by entering [ n ] in the menu.', style='error')

    console.input('\npress [ enter ] to continue > ')

def edit():
    global record

    hr()
    console.print('What class would you like to edit/ delete (note that this is case-sensitive)? (enter [ x ] to cancel)', style='prompt')
    edit_class = console.input('> ')

    if edit_class == 'x':
        menu()
    elif edit_class in record_classes():
        console.print(f'\nWhat value of {edit_class} would you like to edit?', style='prompt')
        console.print('[ 1 ] class name')
        console.print('[ 2 ] letter grade')
        console.print('[ 3 ] class weight')
        console.print('[ 4 ] delete class')
        console.print('[ x ] cancel')
        inpt = console.input('> ')

        if inpt == '1':
            console.print(f'\nWhat would you like to change the name of {edit_class} to?', style='prompt')
            new = console.input('> ')

            existing_classes = record_classes().remove(edit_class)

            if new and len(new) <= 24 and new not in existing_classes:
                ls = record_classes()
                idx = 0
                itr = 0
                while itr < len(ls):
                    if ls[itr] == edit_class:
                        idx = itr
                    itr += 1

                record[idx]['class'] = new

                console.print('\nClass name successfully changed!', style='success')
                console.input('press [ enter ] to continue > ')
                menu()
        elif inpt == 'x':
            menu()
        else:
            console.print('\n[ERROR] Please enter a valid option.', style='error')
            console.input('press [ enter ] to continue > ')
            edit()
    else:
        console.print('\n[ERROR] No matching class name. Please try again.', style='error')
        console.input('\npress [ enter ] to continue > ')
        edit()

def export():
    hr()
    console.print("Would you like to convert your current record to a Excel (.xlsx) file?", style='prompt')
    console.print('[ y ] yes')
    console.print('[ n ] no')

    inpt = console.input('\n> ')

    if inpt == 'y' and len(record) > 0:
        files = listdir()
        if 'records' not in files:
            mkdir('records')

        i = 1
        while exists(f'gpacalc{i}.xlsx'):
            i += 1

        file_name = f'gpacalc{i}.xlsx'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'gpacalc'
        worksheet.sheet_properties.tabColor = "5993f0"

        worksheet.cell(row=1, column=1, value='CLASS')
        worksheet.cell(row=1, column=2, value='GRADE')
        worksheet.cell(row=1, column=3, value='WEIGHT')
        worksheet.cell(row=1, column=4, value='GPA')

        itr = 0
        while itr < len(record):
            worksheet.cell(row=itr+2, column=1, value=record[itr]['class'])
            worksheet.cell(row=itr+2, column=2, value=record[itr]['grade'])
            worksheet.cell(row=itr+2, column=3, value=record[itr]['weight'])
            worksheet.cell(row=itr+2, column=4, value=record[itr]['gpa'])
            itr += 1

        workbook.save(filename=f'records/{file_name}')

        console.print(f"\nSpreadsheet titled '{file_name}' successfully generated in the records folder!", style='success')
        console.print('To open this file, you would need an external tool such as Google Sheets or Microsoft Excel.')
        console.input('press [ enter ] to continue > ')
        menu()
    elif inpt == 'y' and len(record) == 0:
        console.print('\n[ERROR] You must have at least one element in your record to export it as .csv', style='error')
        console.input('press [ enter ] to continue > ')
    elif inpt == 'n':
        menu()
    else:
        console.print('\n[ERROR] Please enter a valid option.', style='error')
        console.input('press [ enter ] to continue > ')

def calculateGPA(weight):
    addGPA = 0
    if weight == 'w': # weighted
        for el in record:
            addGPA += el['gpa']
    elif weight == 'u': # unweighted
        for el in record:
            if el['grade'] == 'a':
                addGPA += 4.0
            elif el['grade'] == 'a-':
                addGPA += 3.667
            elif el['grade'] == 'b+':
                addGPA += 3.333
            elif el['grade'] == 'b':
                addGPA += 3.0
            elif el['grade'] == 'b-':
                addGPA += 2.667
            elif el['grade'] == 'c+':
                addGPA += 2.333
            elif el['grade'] == 'c':
                addGPA += 2.0
            elif el['grade'] == 'c-':
                addGPA += 1.667
            elif el['grade'] == 'd+':
                addGPA += 1.333
            elif el['grade'] == 'd':
                addGPA += 1.0
            elif el['grade'] == 'd-':
                addGPA += 0.667
            elif el['grade'] == 'f':
                addGPA += 0.0
    return round(addGPA / len(record), 3)
    
def record_classes():
    class_names = []
    for el in record:
        class_names.append(f'{el["class"]}')
    return class_names

init()
